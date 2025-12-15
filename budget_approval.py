import pandas as pd
import os
import json
from datetime import datetime
from typing import Dict, List, Tuple, Optional

class BudgetApprovalSystem:
    """
    Budget Approval System for processing and approving budgets created by main.py
    """
    
    def __init__(self):
        self.approval_rules = {
            "max_total_amount": 1000000,  # Maximum total budget amount
            "max_category_percentage": 50,  # Maximum percentage for any single category
            "max_item_percentage": 30,  # Maximum percentage for any single item
            "required_categories": ["Emergency Fund", "Savings"],  # Categories that must be present
            "min_emergency_percentage": 10,  # Minimum percentage for emergency fund
        }
        self.output_dir = "output"
        self.approval_log_file = os.path.join(self.output_dir, "approval_log.json")
    
    def load_budget_from_csv(self, csv_file_path: str) -> Optional[pd.DataFrame]:
        """
        Load budget data from CSV file created by main.py
        """
        try:
            if not os.path.exists(csv_file_path):
                print(f"Error: File '{csv_file_path}' not found.")
                return None
            
            df = pd.read_csv(csv_file_path)
            
            # Validate required columns
            required_columns = ['Category', 'Name', 'Formatted Amount', 'Percentage']
            if not all(col in df.columns for col in required_columns):
                print(f"Error: CSV file must contain columns: {required_columns}")
                return None
            
            # Extract numeric amount from formatted amount
            df['Amount'] = df['Formatted Amount'].str.replace(r'[^\d.]', '', regex=True).astype(float)
            
            print(f"Successfully loaded budget from '{csv_file_path}'")
            print(f"Total items: {len(df)}")
            return df
            
        except Exception as e:
            print(f"Error loading CSV file: {str(e)}")
            return None
    
    def calculate_total_amount(self, df: pd.DataFrame) -> float:
        """Calculate total budget amount"""
        return df['Amount'].sum()
    
    def get_category_summary(self, df: pd.DataFrame) -> pd.DataFrame:
        """Get summary by category"""
        category_summary = df.groupby('Category').agg({
            'Amount': 'sum',
            'Percentage': 'sum',
            'Name': 'count'
        }).rename(columns={'Name': 'Item_Count'})
        
        return category_summary.reset_index()
    
    def check_approval_rules(self, df: pd.DataFrame) -> Tuple[bool, List[str]]:
        """
        Check budget against approval rules
        Returns: (is_approved, list_of_issues)
        """
        issues = []
        
        # Check total amount
        total_amount = self.calculate_total_amount(df)
        if total_amount > self.approval_rules["max_total_amount"]:
            issues.append(f"Total budget ({total_amount:,.0f}) exceeds maximum allowed ({self.approval_rules['max_total_amount']:,.0f})")
        
        # Check category percentages
        category_summary = self.get_category_summary(df)
        for _, row in category_summary.iterrows():
            if row['Percentage'] > self.approval_rules["max_category_percentage"]:
                issues.append(f"Category '{row['Category']}' ({row['Percentage']:.1f}%) exceeds maximum allowed ({self.approval_rules['max_category_percentage']}%)")
        
        # Check individual item percentages
        for _, row in df.iterrows():
            if row['Percentage'] > self.approval_rules["max_item_percentage"]:
                issues.append(f"Item '{row['Name']}' ({row['Percentage']:.1f}%) exceeds maximum allowed ({self.approval_rules['max_item_percentage']}%)")
        
        # Check required categories
        existing_categories = df['Category'].unique()
        for required_cat in self.approval_rules["required_categories"]:
            if required_cat not in existing_categories:
                issues.append(f"Required category '{required_cat}' is missing")
        
        # Check emergency fund minimum
        emergency_percentage = df[df['Category'].str.contains('Emergency', case=False, na=False)]['Percentage'].sum()
        if emergency_percentage < self.approval_rules["min_emergency_percentage"]:
            issues.append(f"Emergency fund ({emergency_percentage:.1f}%) is below minimum required ({self.approval_rules['min_emergency_percentage']}%)")
        
        is_approved = len(issues) == 0
        return is_approved, issues
    
    def display_budget_analysis(self, df: pd.DataFrame):
        """Display comprehensive budget analysis"""
        print("\n" + "="*60)
        print("BUDGET ANALYSIS REPORT")
        print("="*60)
        
        # Total summary
        total_amount = self.calculate_total_amount(df)
        print(f"\nTotal Budget Amount: {total_amount:,.0f}")
        print(f"Total Items: {len(df)}")
        
        # Category breakdown
        print("\nCATEGORY BREAKDOWN:")
        print("-" * 40)
        category_summary = self.get_category_summary(df)
        for _, row in category_summary.iterrows():
            print(f"{row['Category']:<20} | {row['Amount']:>10,.0f} | {row['Percentage']:>6.1f}% | {row['Item_Count']:>3} items")
        
        # Top 5 largest expenses
        print(f"\nTOP 5 LARGEST EXPENSES:")
        print("-" * 40)
        top_expenses = df.nlargest(5, 'Amount')[['Name', 'Category', 'Amount', 'Percentage']]
        for _, row in top_expenses.iterrows():
            print(f"{row['Name']:<25} | {row['Category']:<15} | {row['Amount']:>8,.0f} | {row['Percentage']:>5.1f}%")
    
    def adjust_budget_amounts(self, df: pd.DataFrame) -> Tuple[pd.DataFrame, str]:
        """
        Allow user to adjust specific budget amounts for individual items
        Returns: (modified_dataframe, adjustment_notes)
        """
        print("\n" + "="*70)
        print("INDIVIDUAL BUDGET ITEM ADJUSTMENT")
        print("="*70)
        print("You can modify the approved amount for each budget item individually")
        
        # Display current budget with clear formatting
        print(f"\nCURRENT BUDGET ITEMS (Total: {self.calculate_total_amount(df):,.0f}):")
        print("-" * 70)
        print(f"{'#':<3} {'Item Name':<30} {'Category':<15} {'Current Amount':>15}")
        print("-" * 70)
        
        for idx, row in df.iterrows():
            print(f"{idx+1:2d}. {row['Name']:<30} {row['Category']:<15} {row['Amount']:>15,.0f}")
        
        adjustments = []
        modified_df = df.copy()
        
        print(f"\n{'='*70}")
        print("ADJUSTMENT OPTIONS:")
        print("1. Select specific items by number (e.g., 1,3,5)")
        print("2. Type 'all' to review all items one by one")
        print("3. Type 'done' to finish without changes")
        print("="*70)
        
        while True:
            selection = input("\nEnter your choice: ").strip().lower()
            
            if selection == 'done':
                print("No adjustments made.")
                break
            elif selection == 'all':
                print("\nReviewing all budget items...")
                for idx in range(len(df)):
                    self._adjust_single_item(df, modified_df, idx, adjustments)
                break
            else:
                try:
                    # Parse selected items
                    if ',' in selection:
                        selected_indices = [int(x.strip()) - 1 for x in selection.split(',')]
                    else:
                        selected_indices = [int(selection) - 1]
                    
                    print(f"\nAdjusting {len(selected_indices)} selected item(s)...")
                    for idx in selected_indices:
                        if 0 <= idx < len(df):
                            self._adjust_single_item(df, modified_df, idx, adjustments)
                        else:
                            print(f"❌ Invalid item number: {idx + 1}")
                    break
                            
                except ValueError:
                    print("❌ Invalid input. Please enter item numbers (e.g., 1,3,5), 'all', or 'done'")
        
        if adjustments:
            # Recalculate percentages after adjustments
            total_new_amount = modified_df['Amount'].sum()
            original_total = df['Amount'].sum()
            modified_df['Percentage'] = round((modified_df['Amount'] / total_new_amount) * 100, 2)
            
            # Update formatted amounts (preserve currency from original)
            self._update_formatted_amounts(modified_df, df)
            
            print(f"\n✅ ADJUSTMENT SUMMARY:")
            print(f"   • Items modified: {len(adjustments)}")
            print(f"   • Original total: {original_total:,.0f}")
            print(f"   • New total: {total_new_amount:,.0f}")
            print(f"   • Total change: {total_new_amount - original_total:+,.0f}")
            
            adjustment_notes = "; ".join(adjustments)
        else:
            adjustment_notes = "No adjustments made"
        
        return modified_df, adjustment_notes

    def _adjust_single_item(self, original_df: pd.DataFrame, modified_df: pd.DataFrame, 
                           idx: int, adjustments: List[str]):
        """
        Adjust a single budget item amount
        """
        row = original_df.iloc[idx]
        current_amount = row['Amount']
        
        print(f"\n📋 ITEM {idx+1}: {row['Name']}")
        print(f"   Category: {row['Category']}")
        print(f"   Current Amount: {current_amount:,.0f}")
        print(f"   Current Percentage: {row['Percentage']:.1f}%")
        
        while True:
            action = input(f"   Action - [k]eep current, [c]hange amount, or [r]emove item: ").strip().lower()
            
            if action in ['k', 'keep']:
                print(f"   ✓ Kept: {row['Name']} = {current_amount:,.0f}")
                break
            elif action in ['r', 'remove']:
                # Set amount to 0 to effectively remove the item
                modified_df.loc[idx, 'Amount'] = 0
                modified_df.loc[idx, 'Original_Amount'] = current_amount
                adjustment_note = f"{row['Name']}: {current_amount:,.0f} → REMOVED"
                adjustments.append(adjustment_note)
                print(f"   ❌ Removed: {row['Name']}")
                break
            elif action in ['c', 'change']:
                while True:
                    try:
                        new_amount_input = input(f"   Enter new approved amount: ").strip()
                        new_amount = float(new_amount_input)
                        
                        if new_amount < 0:
                            print("   ❌ Amount cannot be negative. Please try again.")
                            continue
                        
                        # Update the DataFrame
                        modified_df.loc[idx, 'Amount'] = new_amount
                        modified_df.loc[idx, 'Original_Amount'] = current_amount
                        
                        # Record the adjustment
                        if new_amount == 0:
                            adjustment_note = f"{row['Name']}: {current_amount:,.0f} → REMOVED"
                        else:
                            adjustment_note = f"{row['Name']}: {current_amount:,.0f} → {new_amount:,.0f}"
                        adjustments.append(adjustment_note)
                        
                        print(f"   ✓ Updated: {row['Name']} = {new_amount:,.0f}")
                        break
                        
                    except ValueError:
                        print("   ❌ Please enter a valid number.")
                break
            else:
                print("   ❌ Invalid choice. Please enter 'k' (keep), 'c' (change), or 'r' (remove).")

    def _update_formatted_amounts(self, modified_df: pd.DataFrame, original_df: pd.DataFrame):
        """
        Update formatted amounts preserving original currency format
        """
        if len(original_df) > 0:
            # Extract currency format from first item
            original_formatted = original_df.iloc[0]['Formatted Amount']
            currency_part = ''.join([c for c in original_formatted if not c.isdigit() and c not in '.,'])
            
            # Update formatted amounts
            modified_df['Formatted Amount'] = modified_df['Amount'].apply(
                lambda x: f"{currency_part}{x:,.0f}" if currency_part else f"{x:,.0f}"
            )

    def partial_approval_process(self, df: pd.DataFrame, issues: List[str]) -> Tuple[bool, pd.DataFrame, str]:
        """
        Enhanced approval process with individual item adjustment options
        Returns: (is_approved, modified_dataframe, notes)
        """
        print("\n" + "="*70)
        print("BUDGET APPROVAL REVIEW")
        print("="*70)
        
        print("\nISSUES IDENTIFIED:")
        for i, issue in enumerate(issues, 1):
            print(f"  {i}. {issue}")
        
        total_amount = self.calculate_total_amount(df)
        print(f"\nCurrent Total Budget: {total_amount:,.0f}")
        print(f"Number of Items: {len(df)}")
        
        print(f"\n{'='*70}")
        print("APPROVAL OPTIONS:")
        print("1. ✅ Approve full budget as-is (override all rules)")
        print("2. 📝 Adjust individual item amounts and approve")
        print("3. ❌ Reject entire budget")
        print("4. 📊 View detailed budget analysis first")
        print("="*70)
        
        while True:
            choice = input("\nSelect option (1-4): ").strip()
            
            if choice == "1":
                reason = input("\nEnter reason for full approval override: ").strip()
                print(f"\n✅ FULL BUDGET APPROVED (Override)")
                print(f"   Reason: {reason}")
                return True, df, f"Full approval override: {reason}"
                
            elif choice == "2":
                print(f"\n{'='*70}")
                print("INDIVIDUAL ITEM ADJUSTMENT MODE")
                print("="*70)
                print("You can now adjust the approved amount for each budget item individually.")
                print("This allows partial approval where some items may be reduced or removed.")
                
                # Allow user to adjust specific amounts
                modified_df, adjustment_notes = self.adjust_budget_amounts(df)
                
                # Show final summary
                new_total = self.calculate_total_amount(modified_df)
                original_total = self.calculate_total_amount(df)
                
                print(f"\n{'='*70}")
                print("ADJUSTMENT SUMMARY:")
                print(f"  Original Total: {original_total:,.0f}")
                print(f"  Adjusted Total: {new_total:,.0f}")
                print(f"  Net Change: {new_total - original_total:+,.0f}")
                print("="*70)
                
                # Final approval confirmation
                if modified_df.equals(df):
                    print("No changes were made to the budget.")
                    approve_unchanged = input("Approve original budget anyway? (y/n): ").strip().lower()
                    if approve_unchanged in ['y', 'yes']:
                        reason = input("Enter approval reason: ").strip()
                        return True, df, f"Approved without changes: {reason}"
                    else:
                        return False, df, "Rejected - no changes made"
                else:
                    approve_adjusted = input("\nApprove the adjusted budget? (y/n): ").strip().lower()
                    
                    if approve_adjusted in ['y', 'yes']:
                        reason = input("Enter approval reason for adjusted budget: ").strip()
                        final_notes = f"Approved with individual item adjustments: {reason}. Changes: {adjustment_notes}"
                        print(f"\n✅ ADJUSTED BUDGET APPROVED!")
                        return True, modified_df, final_notes
                    else:
                        print("❌ Adjusted budget rejected.")
                        return False, df, f"Rejected after item adjustments. Changes attempted: {adjustment_notes}"
                    
            elif choice == "3":
                reason = input("\nEnter reason for budget rejection: ").strip()
                print(f"\n❌ BUDGET REJECTED")
                print(f"   Reason: {reason}")
                return False, df, f"Budget rejected: {reason}"
                
            elif choice == "4":
                self.display_budget_analysis(df)
                print(f"\n{'='*70}")
                print("Returning to approval options...")
                print("="*70)
                
            else:
                print("❌ Invalid choice. Please enter 1, 2, 3, or 4.")

    def manual_approval_process(self, df: pd.DataFrame, issues: List[str]) -> bool:
        """
        Legacy manual approval process - now redirects to enhanced partial approval
        """
        approved, modified_df, notes = self.partial_approval_process(df, issues)
        return approved
    
    def log_approval_decision(self, csv_file: str, df: pd.DataFrame, approved: bool, issues: List[str], notes: str = ""):
        """Log approval decision to JSON file"""
        
        os.makedirs(self.output_dir, exist_ok=True)
        
        # Load existing log or create new
        approval_log = []
        if os.path.exists(self.approval_log_file):
            try:
                with open(self.approval_log_file, 'r') as f:
                    approval_log = json.load(f)
            except:
                approval_log = []
        
        # Create new log entry
        modifications = []
        if 'Original_Amount' in df.columns:
            for _, row in df.iterrows():
                if pd.notna(row.get('Original_Amount')) and row['Original_Amount'] != row['Amount']:
                    modifications.append({
                        "item": row['Name'],
                        "original_amount": float(row['Original_Amount']),
                        "approved_amount": float(row['Amount'])
                    })
        
        log_entry = {
            "timestamp": datetime.now().isoformat(),
            "csv_file": csv_file,
            "total_amount": float(self.calculate_total_amount(df)),
            "total_items": len(df),
            "categories": df['Category'].unique().tolist(),
            "approved": approved,
            "issues": issues,
            "notes": notes,
            "modifications": modifications,
            "has_modifications": len(modifications) > 0
        }
        
        approval_log.append(log_entry)
        
        # Save updated log
        with open(self.approval_log_file, 'w') as f:
            json.dump(approval_log, f, indent=2)
        
        print(f"Approval decision logged to '{self.approval_log_file}'")
    
    def generate_approval_report(self, csv_file: str, df: pd.DataFrame, approved: bool, issues: List[str]):
        """Generate detailed approval report"""
        
        report_filename = f"approval_report_{datetime.now().strftime('%Y%m%d_%H%M%S')}.txt"
        report_path = os.path.join(self.output_dir, report_filename)
        
        with open(report_path, 'w') as f:
            f.write("BUDGET APPROVAL REPORT\n")
            f.write("=" * 50 + "\n\n")
            f.write(f"Source File: {csv_file}\n")
            f.write(f"Analysis Date: {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}\n")
            f.write(f"Status: {'APPROVED' if approved else 'REJECTED'}\n\n")
            
            # Budget summary
            f.write("BUDGET SUMMARY:\n")
            f.write("-" * 20 + "\n")
            f.write(f"Total Amount: {self.calculate_total_amount(df):,.0f}\n")
            f.write(f"Total Items: {len(df)}\n")
            f.write(f"Categories: {', '.join(df['Category'].unique())}\n\n")
            
            # Issues (if any)
            if issues:
                f.write("ISSUES IDENTIFIED:\n")
                f.write("-" * 20 + "\n")
                for i, issue in enumerate(issues, 1):
                    f.write(f"{i}. {issue}\n")
                f.write("\n")
            
            # Category breakdown
            f.write("CATEGORY BREAKDOWN:\n")
            f.write("-" * 20 + "\n")
            category_summary = self.get_category_summary(df)
            for _, row in category_summary.iterrows():
                f.write(f"{row['Category']}: {row['Amount']:,.0f} ({row['Percentage']:.1f}%) - {row['Item_Count']} items\n")
        
        print(f"Detailed report saved to '{report_path}'")
    
    def process_budget_approval(self, csv_file_path: str):
        """
        Main method to process budget approval with partial approval support
        """
        print("Starting Budget Approval Process...")
        print("=" * 50)
        
        # Load budget data
        df = self.load_budget_from_csv(csv_file_path)
        if df is None:
            return
        
        # Display basic info
        print(f"\nBudget Overview:")
        print(f"File: {csv_file_path}")
        print(f"Total Items: {len(df)}")
        print(f"Total Amount: {self.calculate_total_amount(df):,.0f}")
        print(f"Categories: {', '.join(df['Category'].unique())}")
        
        # Check approval rules
        is_approved, issues = self.check_approval_rules(df)
        modified_df = df.copy()  # Default to original dataframe
        
        if is_approved:
            print("\n✅ BUDGET AUTOMATICALLY APPROVED")
            print("All approval rules satisfied.")
            notes = "Automatic approval - all rules satisfied"
        else:
            print("\n❌ AUTOMATIC APPROVAL FAILED")
            is_approved, modified_df, notes = self.partial_approval_process(df, issues)
        
        # Save modified budget if changes were made
        if not modified_df.equals(df) and is_approved:
            self.display_approved_budget_summary(df, modified_df)
            self.save_approved_budget(csv_file_path, modified_df, notes)
        
        # Log decision and generate report (use modified df for reporting)
        self.log_approval_decision(csv_file_path, modified_df, is_approved, issues, notes)
        self.generate_approval_report(csv_file_path, modified_df, is_approved, issues)
        
        print(f"\n{'='*70}")
        print(f"FINAL STATUS: {'✅ APPROVED' if is_approved else '❌ REJECTED'}")
        if is_approved:
            final_total = self.calculate_total_amount(modified_df)
            original_total = self.calculate_total_amount(df)
            if final_total != original_total:
                print(f"Original Amount: {original_total:,.0f}")
                print(f"Approved Amount: {final_total:,.0f}")
                print(f"Net Change: {final_total - original_total:+,.0f}")
            else:
                print(f"Approved Amount: {final_total:,.0f}")
        print("="*70)

    def display_approved_budget_summary(self, original_df: pd.DataFrame, approved_df: pd.DataFrame):
        """
        Display a clear summary of the approved budget with changes highlighted
        """
        print(f"\n{'='*80}")
        print("APPROVED BUDGET SUMMARY")
        print("="*80)
        
        original_total = self.calculate_total_amount(original_df)
        approved_total = self.calculate_total_amount(approved_df)
        
        print(f"Original Total:  {original_total:>15,.0f}")
        print(f"Approved Total:  {approved_total:>15,.0f}")
        print(f"Net Change:      {approved_total - original_total:>+15,.0f}")
        print("-" * 80)
        
        # Show item-by-item comparison
        print(f"{'Item':<30} {'Category':<15} {'Original':>12} {'Approved':>12} {'Change':>12}")
        print("-" * 80)
        
        for idx, row in approved_df.iterrows():
            original_amount = row.get('Original_Amount', row['Amount'])
            approved_amount = row['Amount']
            change = approved_amount - original_amount
            
            change_indicator = ""
            if change > 0:
                change_indicator = "↗"
            elif change < 0:
                change_indicator = "↘"
            elif approved_amount == 0:
                change_indicator = "❌"
            else:
                change_indicator = "✓"
            
            print(f"{row['Name']:<30} {row['Category']:<15} {original_amount:>12,.0f} "
                  f"{approved_amount:>12,.0f} {change:>+11,.0f} {change_indicator}")
        
        print("="*80)

    def save_approved_budget(self, original_file_path: str, approved_df: pd.DataFrame, notes: str):
        """
        Save the approved budget with modifications to both CSV and Excel with neat column titles
        """
        # Generate approved budget filename base
        base_name = os.path.splitext(os.path.basename(original_file_path))[0]
        timestamp = datetime.now().strftime('%Y%m%d_%H%M%S')
        
        # Extract currency from the original formatted amount
        currency_symbol = ""
        if len(approved_df) > 0 and 'Formatted Amount' in approved_df.columns:
            original_formatted = approved_df.iloc[0]['Formatted Amount']
            currency_symbol = ''.join([c for c in original_formatted if not c.isdigit() and c not in '.,'])
            
            # Update formatted amounts
            approved_df['Formatted Amount'] = approved_df['Amount'].apply(
                lambda x: f"{currency_symbol}{x:,.0f}" if currency_symbol else f"{x:,.0f}"
            )
        
        # 1. Save CSV version
        approved_csv_filename = f"{base_name}_APPROVED_{timestamp}.csv"
        approved_csv_path = os.path.join(self.output_dir, approved_csv_filename)
        approved_df.to_csv(approved_csv_path, index=False, 
                          columns=['Category', 'Name', 'Formatted Amount', 'Percentage'])
        print(f"\n✅ Approved budget (CSV) saved to: '{approved_csv_path}'")
        
        # 2. Save Excel version with professional formatting
        try:
            approved_excel_filename = f"{base_name}_APPROVED_REPORT_{timestamp}.xlsx"
            approved_excel_path = os.path.join(self.output_dir, approved_excel_filename)
            
            # Create neat column mapping for professional reports
            column_mapping = {
                'Category': 'Budget Category',
                'Name': 'Item Description',
                'Formatted Amount': f'Approved Amount',
                'Percentage': 'Percentage (%)',
                'Original_Amount': 'Original Amount (Numeric)'
            }
            
            # Prepare export dataframe
            export_cols = ['Category', 'Name', 'Formatted Amount', 'Percentage']
            if 'Original_Amount' in approved_df.columns:
                # Add comparison column if there were modifications
                approved_df['Change'] = approved_df['Amount'] - approved_df['Original_Amount'].fillna(approved_df['Amount'])
                approved_df['Change_Formatted'] = approved_df['Change'].apply(
                    lambda x: f"{'+' if x > 0 else ''}{currency_symbol}{x:,.0f}" if x != 0 else "-"
                )
                export_cols.extend(['Change_Formatted'])
                column_mapping['Change_Formatted'] = 'Amount Change'
            
            df_export = approved_df[export_cols].copy()
            df_export = df_export.rename(columns=column_mapping)
            
            # Export with formatting using openpyxl
            with pd.ExcelWriter(approved_excel_path, engine='openpyxl') as writer:
                df_export.to_excel(writer, sheet_name='Approved Budget', index=False)
                
                # Get workbook and worksheet for formatting
                workbook = writer.book
                worksheet = writer.sheets['Approved Budget']
                
                # Import styling modules
                from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
                
                # Header formatting
                header_fill = PatternFill(start_color='366092', end_color='366092', fill_type='solid')
                header_font = Font(bold=True, color='FFFFFF', size=11)
                
                for col_idx, col in enumerate(df_export.columns, 1):
                    # Format header
                    cell = worksheet.cell(row=1, column=col_idx)
                    cell.fill = header_fill
                    cell.font = header_font
                    cell.alignment = Alignment(horizontal='center', vertical='center')
                    
                    # Auto-adjust column width
                    max_length = max(
                        df_export[col].astype(str).apply(len).max(),
                        len(col)
                    ) + 2
                    column_letter = cell.column_letter
                    worksheet.column_dimensions[column_letter].width = min(max_length, 50)
                
                # Add borders to all cells
                thin_border = Border(
                    left=Side(style='thin'),
                    right=Side(style='thin'),
                    top=Side(style='thin'),
                    bottom=Side(style='thin')
                )
                
                for row in worksheet.iter_rows(min_row=1, max_row=len(df_export)+1, 
                                              min_col=1, max_col=len(df_export.columns)):
                    for cell in row:
                        cell.border = thin_border
                        if cell.row > 1:  # Data rows
                            cell.alignment = Alignment(horizontal='left', vertical='center')
                
                # Add summary information at the bottom
                summary_row = len(df_export) + 3
                worksheet.cell(row=summary_row, column=1, value="Approval Summary")
                worksheet.cell(row=summary_row, column=1).font = Font(bold=True, size=12)
                
                worksheet.cell(row=summary_row+1, column=1, value="Total Approved Amount:")
                worksheet.cell(row=summary_row+1, column=2, 
                             value=f"{currency_symbol}{self.calculate_total_amount(approved_df):,.0f}")
                
                worksheet.cell(row=summary_row+2, column=1, value="Approval Date:")
                worksheet.cell(row=summary_row+2, column=2, 
                             value=datetime.now().strftime('%Y-%m-%d %H:%M:%S'))
                
                if 'Original_Amount' in approved_df.columns:
                    original_total = approved_df['Original_Amount'].fillna(approved_df['Amount']).sum()
                    approved_total = self.calculate_total_amount(approved_df)
                    change = approved_total - original_total
                    
                    worksheet.cell(row=summary_row+3, column=1, value="Original Amount:")
                    worksheet.cell(row=summary_row+3, column=2, 
                                 value=f"{currency_symbol}{original_total:,.0f}")
                    
                    worksheet.cell(row=summary_row+4, column=1, value="Net Change:")
                    worksheet.cell(row=summary_row+4, column=2, 
                                 value=f"{'+' if change > 0 else ''}{currency_symbol}{change:,.0f}")
            
            print(f"✅ Approved budget (Excel Report) saved to: '{approved_excel_path}'")
            print(f"   Format: Professional report with neat column titles and formatting")
            
        except ImportError:
            print("\n⚠️ openpyxl not installed. Excel export skipped. Install with: pip install openpyxl")
        except Exception as e:
            print(f"\n⚠️ Excel export failed: {str(e)}. CSV version is available.")
        
        # 3. Save approval summary (text file)
        summary_filename = f"{base_name}_APPROVAL_SUMMARY_{timestamp}.txt"
        summary_file_path = os.path.join(self.output_dir, summary_filename)
        
        with open(summary_file_path, 'w') as f:
            f.write("BUDGET APPROVAL SUMMARY\n")
            f.write("=" * 50 + "\n\n")
            f.write(f"Original File: {original_file_path}\n")
            f.write(f"Approved CSV: {approved_csv_path}\n")
            f.write(f"Approval Date: {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}\n")
            f.write(f"Total Approved Amount: {self.calculate_total_amount(approved_df):,.0f}\n\n")
            f.write(f"Approval Notes: {notes}\n\n")
            
            # Show changes if any
            if 'Original_Amount' in approved_df.columns:
                f.write("AMOUNT MODIFICATIONS:\n")
                f.write("-" * 30 + "\n")
                for _, row in approved_df.iterrows():
                    if pd.notna(row.get('Original_Amount')) and row['Original_Amount'] != row['Amount']:
                        f.write(f"{row['Name']}: {row['Original_Amount']:,.0f} → {row['Amount']:,.0f}\n")
        
        print(f"✅ Approval summary saved to: '{summary_file_path}'")


def main():
    """Main function to run budget approval system"""
    approval_system = BudgetApprovalSystem()
    
    print("Budget Approval System")
    print("=" * 30)
    
    # List available CSV files
    output_dir = "output"
    if not os.path.exists(output_dir):
        print(f"No '{output_dir}' directory found. Please run main.py first to create budgets.")
        return
    
    csv_files = [f for f in os.listdir(output_dir) if f.endswith('.csv')]
    
    if not csv_files:
        print(f"No CSV files found in '{output_dir}' directory.")
        print("Please run main.py first to create a budget.")
        return
    
    print("\nAvailable budget files:")
    for i, file in enumerate(csv_files, 1):
        print(f"{i}. {file}")
    
    # Get user choice
    while True:
        try:
            choice = input(f"\nSelect a file to process (1-{len(csv_files)}) or 'q' to quit: ").strip()
            if choice.lower() == 'q':
                print("Exiting...")
                return
            
            file_index = int(choice) - 1
            if 0 <= file_index < len(csv_files):
                selected_file = os.path.join(output_dir, csv_files[file_index])
                approval_system.process_budget_approval(selected_file)
                break
            else:
                print("Invalid choice. Please try again.")
        except ValueError:
            print("Please enter a valid number or 'q' to quit.")


if __name__ == "__main__":
    main()
