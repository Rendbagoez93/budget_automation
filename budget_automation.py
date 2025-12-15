import pandas as pd
import os
import sys
from datetime import datetime

# Import budget approval system
try:
    from budget_approval import BudgetApprovalSystem
    APPROVAL_SYSTEM_AVAILABLE = True
except ImportError:
    print("Warning: budget_approval.py not found. Approval system will be disabled.")
    APPROVAL_SYSTEM_AVAILABLE = False

def format_currency(amount, currency):
    """Format amount with currency symbol and comma separator"""
    symbols = {
        "IDR": "Rp",
        "USD": "$",
        "EUR": "€",
        "JPY": "¥"
        # Add more if needed
    }
    symbol = symbols.get(currency.upper(), currency + " ")
    return f"{symbol}{float(amount):,.0f}"

def create_budget(currency, entries):
    """
    Create a budget DataFrame with percentage breakdown and formatted currency amounts.
    :param currency: str, currency code (e.g. 'IDR', 'USD')
    :param entries: list of dicts with 'Category', 'Name', and 'Amount'
    :return: pandas DataFrame
    """
    df = pd.DataFrame(entries)
    df['Amount'] = pd.to_numeric(df['Amount'], errors='coerce')
    
    # Check for invalid amounts (NaN values)
    if df['Amount'].isna().any():
        invalid_entries = df[df['Amount'].isna()]['Name'].tolist()
        print(f"Warning: Invalid amounts detected for: {', '.join(invalid_entries)}")
        print("These entries will be excluded from calculations.")
        df = df.dropna(subset=['Amount'])
    
    if df.empty:
        print("Error: No valid budget entries found.")
        return pd.DataFrame()
    
    total = df['Amount'].sum()
    df['Percentage'] = round((df['Amount'] / total) * 100, 2)
    df['Formatted Amount'] = df.apply(lambda row: format_currency(row['Amount'], currency), axis=1)

    print(f"\nTotal Budget: {format_currency(total, currency)}\n")
    print(f"Created on: {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}\n")
    print("Budget Breakdown:")
    print("-" * 60)
    print(df[['Category', 'Name', 'Formatted Amount', 'Percentage']])

    return df


def save_budget_to_csv(budget_df, currency):
    """
    Save budget DataFrame to CSV file with enhanced naming
    """
    os.makedirs("output", exist_ok=True)
    
    # Generate default filename with timestamp
    timestamp = datetime.now().strftime('%Y%m%d_%H%M%S')
    default_filename = f"budget_{currency.lower()}_{timestamp}.csv"
    
    print(f"\nDefault filename: {default_filename}")
    custom_name = input("Enter custom filename (or press Enter to use default): ").strip()
    
    if custom_name:
        file_name = custom_name
        if not file_name.endswith('.csv'):
            file_name += '.csv'
    else:
        file_name = default_filename
    
    file_path = os.path.join("output", file_name)
    budget_df.to_csv(file_path, index=False, columns=['Category', 'Name', 'Formatted Amount', 'Percentage'])
    print(f"\nBudget saved to '{file_path}' successfully!")
    
    return file_path


def export_to_excel(budget_df, currency, include_approval_columns=False):
    """
    Export budget to Excel with professional formatting and neat column titles
    Perfect for reports and presentations
    """
    os.makedirs("output", exist_ok=True)
    
    # Generate default filename with timestamp
    timestamp = datetime.now().strftime('%Y%m%d_%H%M%S')
    default_filename = f"budget_report_{currency.lower()}_{timestamp}.xlsx"
    
    print(f"\n{'='*60}")
    print("EXPORT TO EXCEL (Professional Report Format)")
    print("='*60}")
    print(f"Default filename: {default_filename}")
    custom_name = input("Enter custom filename (or press Enter to use default): ").strip()
    
    if custom_name:
        file_name = custom_name
        if not file_name.endswith('.xlsx'):
            file_name += '.xlsx'
    else:
        file_name = default_filename
    
    file_path = os.path.join("output", file_name)
    
    # Create neat column mapping for professional reports
    column_mapping = {
        'Category': 'Budget Category',
        'Name': 'Item Description',
        'Formatted Amount': f'Amount ({currency})',
        'Percentage': 'Percentage (%)',
        'Amount': 'Numeric Amount'
    }
    
    # Add approval columns if requested
    if include_approval_columns:
        column_mapping.update({
            'Approval_Status': 'Approval Status',
            'Approved_By': 'Approved By',
            'Approval_Date': 'Approval Date',
            'Comments': 'Comments'
        })
    
    # Select and rename columns
    if include_approval_columns:
        export_columns = ['Category', 'Name', 'Formatted Amount', 'Percentage']
        # Add approval columns if they exist
        for col in ['Approval_Status', 'Approved_By', 'Approval_Date', 'Comments']:
            if col in budget_df.columns:
                export_columns.append(col)
    else:
        export_columns = ['Category', 'Name', 'Formatted Amount', 'Percentage']
    
    df_export = budget_df[export_columns].copy()
    df_export = df_export.rename(columns=column_mapping)
    
    # Export with formatting
    try:
        with pd.ExcelWriter(file_path, engine='openpyxl') as writer:
            df_export.to_excel(writer, sheet_name='Budget Report', index=False)
            
            # Get worksheet for formatting
            worksheet = writer.sheets['Budget Report']
            
            # Auto-adjust column widths
            from openpyxl.styles import Font, PatternFill, Alignment
            
            # Header formatting
            header_fill = PatternFill(start_color='366092', end_color='366092', fill_type='solid')
            header_font = Font(bold=True, color='FFFFFF', size=11)
            
            for col_idx, col in enumerate(df_export.columns, 1):
                # Format header
                cell = worksheet.cell(row=1, column=col_idx)
                cell.fill = header_fill
                cell.font = header_font
                cell.alignment = Alignment(horizontal='center', vertical='center')
                
                # Auto-adjust column width
                max_length = max(
                    df_export[col].astype(str).apply(len).max(),
                    len(col)
                ) + 2
                column_letter = worksheet.cell(row=1, column=col_idx).column_letter
                worksheet.column_dimensions[column_letter].width = min(max_length, 50)
            
            # Add border to all cells
            from openpyxl.styles import Border, Side
            thin_border = Border(
                left=Side(style='thin'),
                right=Side(style='thin'),
                top=Side(style='thin'),
                bottom=Side(style='thin')
            )
            
            for row in worksheet.iter_rows(min_row=1, max_row=len(df_export)+1, 
                                          min_col=1, max_col=len(df_export.columns)):
                for cell in row:
                    cell.border = thin_border
                    if cell.row > 1:  # Data rows
                        cell.alignment = Alignment(horizontal='left', vertical='center')
        
        print("\n✅ Excel report exported successfully!")
        print("   Location: '{}'" .format(file_path))
        print("   Format: Professional report with formatted columns")
        return file_path
        
    except ImportError:
        print("\n⚠️ openpyxl not installed. Installing now...")
        import subprocess
        subprocess.check_call([sys.executable, "-m", "pip", "install", "openpyxl"])
        print("✅ openpyxl installed. Please run export again.")
        return None
    except Exception as e:
        print(f"\n❌ Error exporting to Excel: {str(e)}")
        print("Falling back to CSV export...")
        return None


def run_approval_process(csv_file_path):
    """
    Run the budget approval process if available
    """
    if not APPROVAL_SYSTEM_AVAILABLE:
        print("\nApproval system not available. Please ensure budget_approval.py exists.")
        return
    
    print("\n" + "="*60)
    print("STARTING BUDGET APPROVAL PROCESS")
    print("="*60)
    
    try:
        approval_system = BudgetApprovalSystem()
        approval_system.process_budget_approval(csv_file_path)
    except Exception as e:
        print(f"Error running approval process: {str(e)}")


def main():
    """
    Main budget automation workflow
    """
    print("="*60)
    print("BUDGET AUTOMATION SYSTEM")
    print("="*60)
    print("Create and optionally approve your budget")
    print()    
    # Get currency from user
    valid_currencies = ["IDR", "USD", "EUR", "JPY"]
    while True:
        currency_input = input("Enter currency (IDR, USD, EUR, JPY): ").upper()
        if currency_input in valid_currencies:
            break
        print(f"Invalid currency. Please choose from: {', '.join(valid_currencies)}")

    # Collect budget entries
    budget_items = []
    print(f"\nEnter budget items (currency: {currency_input})")
    print("Type 'done' when finished entering items")
    print("-" * 40)
    
    while True:
        category = input("Enter budget category (or type 'done' to finish): ").strip()
        if category.lower() == 'done':
            break
        if not category:
            print("Category cannot be empty. Please try again.")
            continue
        
        name = input("Enter budget name: ").strip()
        if not name:
            print("Budget name cannot be empty. Please try again.")
            continue
        
        while True:
            try:
                amount = input(f"Enter amount for '{name}': ").strip()
                if not amount:
                    print("Amount cannot be empty. Please try again.")
                    continue
                # Test if it's a valid number
                float(amount)
                break
            except ValueError:
                print("Please enter a valid number.")
        
        budget_items.append({'Category': category, 'Name': name, 'Amount': amount})
        print(f"✓ Added: {name} - {format_currency(float(amount), currency_input)}")

    if not budget_items:
        print("No budget items entered. Exiting.")
        return

    # Generate and display budget
    budget_df = create_budget(currency_input, budget_items)

    # Only proceed with export if we have valid data
    if not budget_df.empty:
        # Ask user for export format preference
        print("\n" + "="*60)
        print("EXPORT OPTIONS:")
        print("="*60)
        print("1. CSV (simple text format)")
        print("2. Excel (formatted report with professional column names)")
        print("3. Both formats")
        
        export_choice = input("\nSelect export format (1/2/3): ").strip()
        
        csv_file_path = None
        excel_file_path = None
        
        if export_choice in ['1', '3']:
            csv_file_path = save_budget_to_csv(budget_df, currency_input)
        
        if export_choice in ['2', '3']:
            excel_file_path = export_to_excel(budget_df, currency_input)
            
        if export_choice not in ['1', '2', '3']:
            print("Invalid choice. Defaulting to CSV export...")
            csv_file_path = save_budget_to_csv(budget_df, currency_input)
        
        # Use CSV path for approval process (if Excel only, save temp CSV)
        if csv_file_path is None and excel_file_path:
            timestamp = datetime.now().strftime('%Y%m%d_%H%M%S')
            csv_file_path = os.path.join("output", f"temp_budget_{timestamp}.csv")
            budget_df.to_csv(csv_file_path, index=False, 
                           columns=['Category', 'Name', 'Formatted Amount', 'Percentage'])
        
        # Ask if user wants to run approval process
        if APPROVAL_SYSTEM_AVAILABLE and csv_file_path:
            run_approval = input("\nWould you like to run the approval process? (y/n): ").strip().lower()
            if run_approval in ['y', 'yes']:
                run_approval_process(csv_file_path)
            else:
                print("Budget saved. You can run approval later using budget_approval.py")
        else:
            print("Budget creation completed.")
    else:
        print("No valid budget data to save.")

if __name__ == "__main__":
    main()