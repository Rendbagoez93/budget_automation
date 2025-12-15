#!/usr/bin/env python3
"""
Comprehensive Budget Creation System
Advanced budget creation with templates, quarterly breakdown, and analysis
Integrates seamlessly with the existing workflow
"""

import os
import sys
import pandas as pd
from datetime import datetime
from budget_templates import (
    BudgetTemplates, 
    QuarterlyBudget, 
    BudgetAnalyzer, 
    BudgetComparison
)

# Import from existing budget_automation for compatibility
try:
    from budget_automation import format_currency, export_to_excel, APPROVAL_SYSTEM_AVAILABLE
    if APPROVAL_SYSTEM_AVAILABLE:
        from budget_approval import BudgetApprovalSystem
except ImportError:
    print("Warning: Could not import budget_automation modules")
    APPROVAL_SYSTEM_AVAILABLE = False


class ComprehensiveBudgetCreator:
    """Comprehensive budget creation with advanced features"""
    
    def __init__(self):
        self.currency = "USD"
        self.budget_items = []
        self.budget_df = None
        self.budget_type = None
        self.output_dir = "output"
        os.makedirs(self.output_dir, exist_ok=True)
    
    def select_currency(self):
        """Select currency for budget"""
        valid_currencies = ["IDR", "USD", "EUR", "JPY", "GBP", "AUD", "CAD", "SGD"]
        
        print("\n" + "="*60)
        print("SELECT CURRENCY")
        print("="*60)
        print("Available currencies:")
        for i, curr in enumerate(valid_currencies, 1):
            print(f"{i}. {curr}")
        
        while True:
            choice = input(f"\nEnter currency code or number (1-{len(valid_currencies)}): ").strip().upper()
            
            if choice.isdigit():
                idx = int(choice) - 1
                if 0 <= idx < len(valid_currencies):
                    self.currency = valid_currencies[idx]
                    break
            elif choice in valid_currencies:
                self.currency = choice
                break
            
            print(f"Invalid input. Please choose from: {', '.join(valid_currencies)}")
        
        print(f"\n✓ Currency set to: {self.currency}")
    
    def select_budget_type(self):
        """Select budget type/template"""
        print("\n" + "="*60)
        print("SELECT BUDGET TYPE")
        print("="*60)
        
        templates = BudgetTemplates.list_available_templates()
        template_keys = list(templates.keys())
        
        for i, (key, description) in enumerate(templates.items(), 1):
            print(f"{i}. {key.upper()}: {description}")
        print(f"{len(templates) + 1}. CUSTOM: Start from scratch")
        
        while True:
            choice = input(f"\nSelect budget type (1-{len(templates) + 1}): ").strip()
            
            if choice.isdigit():
                idx = int(choice) - 1
                if 0 <= idx < len(templates):
                    self.budget_type = template_keys[idx]
                    print(f"\n✓ Selected: {self.budget_type.upper()} budget template")
                    return True
                elif idx == len(templates):
                    self.budget_type = 'custom'
                    print("\n✓ Selected: CUSTOM budget (no template)")
                    return False
            
            print("Invalid choice. Please try again.")
    
    def use_template(self):
        """Load and fill template"""
        template_items = BudgetTemplates.get_template_by_name(self.budget_type)
        
        if not template_items:
            print("Error: Could not load template")
            return False
        
        print(f"\n✓ Loaded {len(template_items)} template items")
        print("\n" + "="*60)
        print("FILL TEMPLATE")
        print("="*60)
        print("Enter amounts for each item (press Enter to skip with $0)")
        print("You can add custom items after completing the template")
        print("-" * 60)
        
        filled_items = []
        
        current_category = None
        for item in template_items:
            category = item['Category']
            name = item['Name']
            priority = item.get('Priority', 'Medium')
            
            # Print category header
            if category != current_category:
                print(f"\n[{category.upper()}]")
                current_category = category
            
            # Get amount
            while True:
                amount_input = input(f"  {name} [{priority}]: {self.currency} ").strip()
                
                if not amount_input:
                    amount = 0
                    break
                
                try:
                    amount = float(amount_input)
                    if amount < 0:
                        print("    ⚠️  Amount cannot be negative")
                        continue
                    break
                except ValueError:
                    print("    ⚠️  Please enter a valid number")
            
            if amount > 0:
                filled_items.append({
                    'Category': category,
                    'Name': name,
                    'Amount': amount,
                    'Priority': priority
                })
                print(f"    ✓ Added: {format_currency(amount, self.currency)}")
        
        self.budget_items = filled_items
        
        # Show summary
        if filled_items:
            total = sum(item['Amount'] for item in filled_items)
            print(f"\n✓ Template filled: {len(filled_items)} items, Total: {format_currency(total, self.currency)}")
            
            # Option to add more items
            add_more = input("\nWould you like to add custom items? (y/n): ").strip().lower()
            if add_more in ['y', 'yes']:
                self.add_custom_items()
            
            return True
        else:
            print("\n⚠️  No items filled from template")
            return False
    
    def add_custom_items(self):
        """Add custom budget items"""
        print("\n" + "="*60)
        print("ADD CUSTOM ITEMS")
        print("="*60)
        print("Type 'done' when finished")
        print("-" * 60)
        
        while True:
            category = input("\nCategory (or 'done'): ").strip()
            if category.lower() == 'done':
                break
            if not category:
                print("⚠️  Category cannot be empty")
                continue
            
            name = input("Item name: ").strip()
            if not name:
                print("⚠️  Name cannot be empty")
                continue
            
            while True:
                try:
                    amount = float(input(f"Amount ({self.currency}): ").strip())
                    if amount < 0:
                        print("⚠️  Amount cannot be negative")
                        continue
                    break
                except ValueError:
                    print("⚠️  Please enter a valid number")
            
            # Optional priority
            priority_input = input("Priority (High/Medium/Low, default: Medium): ").strip()
            priority = priority_input if priority_input in ['High', 'Medium', 'Low'] else 'Medium'
            
            self.budget_items.append({
                'Category': category,
                'Name': name,
                'Amount': amount,
                'Priority': priority
            })
            
            print(f"✓ Added: {name} - {format_currency(amount, self.currency)}")
    
    def create_budget_dataframe(self):
        """Convert budget items to DataFrame with calculations"""
        if not self.budget_items:
            print("No budget items to process")
            return None
        
        df = pd.DataFrame(self.budget_items)
        df['Amount'] = pd.to_numeric(df['Amount'], errors='coerce').fillna(0)
        
        # Calculate percentages
        total = df['Amount'].sum()
        if total > 0:
            df['Percentage'] = round((df['Amount'] / total) * 100, 2)
        else:
            df['Percentage'] = 0
        
        # Format currency amounts
        df['Formatted Amount'] = df.apply(
            lambda row: format_currency(row['Amount'], self.currency), 
            axis=1
        )
        
        self.budget_df = df
        return df
    
    def display_budget_summary(self):
        """Display comprehensive budget summary"""
        if self.budget_df is None:
            return
        
        print("\n" + "="*70)
        print("BUDGET SUMMARY")
        print("="*70)
        
        total = self.budget_df['Amount'].sum()
        print(f"\nTotal Budget: {format_currency(total, self.currency)}")
        print(f"Currency: {self.currency}")
        print(f"Budget Type: {self.budget_type.upper() if self.budget_type else 'Custom'}")
        print(f"Created: {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}")
        print(f"Total Items: {len(self.budget_df)}")
        print(f"Total Categories: {self.budget_df['Category'].nunique()}")
        
        # Category breakdown
        print("\n" + "-"*70)
        print("CATEGORY BREAKDOWN")
        print("-"*70)
        category_summary = self.budget_df.groupby('Category').agg({
            'Amount': 'sum',
            'Percentage': 'sum',
            'Name': 'count'
        }).rename(columns={'Name': 'Items'}).sort_values('Amount', ascending=False)
        
        print(f"{'Category':<25} {'Amount':>15} {'Percentage':>12} {'Items':>8}")
        print("-"*70)
        for idx, row in category_summary.iterrows():
            print(f"{idx:<25} {format_currency(row['Amount'], self.currency):>15} "
                  f"{row['Percentage']:>11.1f}% {int(row['Items']):>8}")
        
        # Priority breakdown
        if 'Priority' in self.budget_df.columns:
            print("\n" + "-"*70)
            print("PRIORITY BREAKDOWN")
            print("-"*70)
            priority_summary = self.budget_df.groupby('Priority').agg({
                'Amount': 'sum',
                'Percentage': 'sum'
            }).sort_values('Amount', ascending=False)
            
            for idx, row in priority_summary.iterrows():
                print(f"{idx:<10} {format_currency(row['Amount'], self.currency):>15} ({row['Percentage']:.1f}%)")
        
        # Top 5 expenses
        print("\n" + "-"*70)
        print("TOP 5 LARGEST EXPENSES")
        print("-"*70)
        top_5 = self.budget_df.nlargest(5, 'Amount')
        for idx, row in top_5.iterrows():
            print(f"{row['Name']:<30} {row['Category']:<20} {format_currency(row['Amount'], self.currency):>15}")
        
        # Recommendations
        print("\n" + "-"*70)
        print("RECOMMENDATIONS")
        print("-"*70)
        recommendations = BudgetAnalyzer.get_recommendations(
            self.budget_df, 
            self.budget_type if self.budget_type != 'custom' else 'general'
        )
        for rec in recommendations:
            print(f"  {rec}")
        
        print("="*70)
    
    def create_quarterly_breakdown(self):
        """Create quarterly budget breakdown"""
        if self.budget_df is None or len(self.budget_df) == 0:
            return None
        
        print("\n" + "="*60)
        print("QUARTERLY BREAKDOWN")
        print("="*60)
        print("Select distribution method:")
        print("1. Equal - Distribute evenly across quarters (25% each)")
        print("2. Weighted - Business standard (Q1:20%, Q2:25%, Q3:25%, Q4:30%)")
        print("3. Seasonal - Retail/Consumer (Q1:15%, Q2:20%, Q3:25%, Q4:40%)")
        print("4. Skip - No quarterly breakdown")
        
        while True:
            choice = input("\nSelect option (1-4): ").strip()
            
            if choice == '1':
                distribution = 'equal'
                break
            elif choice == '2':
                distribution = 'weighted'
                break
            elif choice == '3':
                distribution = 'seasonal'
                break
            elif choice == '4':
                return None
            else:
                print("Invalid choice. Please select 1-4")
        
        # Create quarterly breakdown
        items = self.budget_df[['Category', 'Name', 'Amount']].to_dict('records')
        quarterly_df = QuarterlyBudget.create_quarterly_breakdown(items, distribution)
        
        # Add other columns
        if 'Priority' in self.budget_df.columns:
            quarterly_df['Priority'] = self.budget_df['Priority']
        
        # Calculate quarterly totals
        q1_total = quarterly_df['Q1'].sum()
        q2_total = quarterly_df['Q2'].sum()
        q3_total = quarterly_df['Q3'].sum()
        q4_total = quarterly_df['Q4'].sum()
        
        print(f"\n✓ Quarterly breakdown created:")
        print(f"  Q1: {format_currency(q1_total, self.currency)}")
        print(f"  Q2: {format_currency(q2_total, self.currency)}")
        print(f"  Q3: {format_currency(q3_total, self.currency)}")
        print(f"  Q4: {format_currency(q4_total, self.currency)}")
        
        return quarterly_df
    
    def save_budget(self, include_quarterly=False):
        """Save budget to files"""
        if self.budget_df is None or len(self.budget_df) == 0:
            print("No budget data to save")
            return None
        
        # Generate filename
        timestamp = datetime.now().strftime('%Y%m%d_%H%M%S')
        budget_type_str = self.budget_type if self.budget_type else 'custom'
        base_filename = f"budget_{budget_type_str}_{self.currency.lower()}_{timestamp}"
        
        print("\n" + "="*60)
        print("SAVE BUDGET")
        print("="*60)
        print(f"Default filename: {base_filename}")
        custom_name = input("Enter custom name (or press Enter to use default): ").strip()
        
        if custom_name:
            base_filename = custom_name
        
        # Ask for format
        print("\nSelect export format:")
        print("1. CSV only")
        print("2. Excel only (with formatting)")
        print("3. Both CSV and Excel")
        
        format_choice = input("Select format (1-3): ").strip()
        
        saved_files = []
        
        # Prepare export DataFrame
        export_cols = ['Category', 'Name', 'Formatted Amount', 'Percentage']
        if 'Priority' in self.budget_df.columns:
            export_cols.insert(2, 'Priority')
        
        export_df = self.budget_df[export_cols].copy()
        
        # Save CSV
        if format_choice in ['1', '3']:
            csv_filename = f"{base_filename}.csv"
            csv_path = os.path.join(self.output_dir, csv_filename)
            export_df.to_csv(csv_path, index=False)
            print(f"\n✓ CSV saved: {csv_path}")
            saved_files.append(csv_path)
        
        # Save Excel
        if format_choice in ['2', '3']:
            excel_filename = f"{base_filename}_report.xlsx"
            excel_path = os.path.join(self.output_dir, excel_filename)
            
            try:
                # Prepare column mapping for professional names
                column_mapping = {
                    'Category': 'Budget Category',
                    'Name': 'Item Description',
                    'Priority': 'Priority Level',
                    'Formatted Amount': f'Amount ({self.currency})',
                    'Percentage': 'Percentage (%)'
                }
                
                df_export = export_df.rename(columns=column_mapping)
                
                with pd.ExcelWriter(excel_path, engine='openpyxl') as writer:
                    df_export.to_excel(writer, sheet_name='Budget', index=False)
                    
                    # Get worksheet for formatting
                    worksheet = writer.sheets['Budget']
                    
                    # Import styling
                    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
                    
                    # Header formatting
                    header_fill = PatternFill(start_color='366092', end_color='366092', fill_type='solid')
                    header_font = Font(bold=True, color='FFFFFF', size=11)
                    
                    for col_idx, col in enumerate(df_export.columns, 1):
                        cell = worksheet.cell(row=1, column=col_idx)
                        cell.fill = header_fill
                        cell.font = header_font
                        cell.alignment = Alignment(horizontal='center', vertical='center')
                        
                        # Auto-adjust width
                        max_length = max(
                            df_export[col].astype(str).apply(len).max(),
                            len(col)
                        ) + 2
                        column_letter = cell.column_letter
                        worksheet.column_dimensions[column_letter].width = min(max_length, 50)
                    
                    # Add borders
                    thin_border = Border(
                        left=Side(style='thin'),
                        right=Side(style='thin'),
                        top=Side(style='thin'),
                        bottom=Side(style='thin')
                    )
                    
                    for row in worksheet.iter_rows(min_row=1, max_row=len(df_export)+1,
                                                   min_col=1, max_col=len(df_export.columns)):
                        for cell in row:
                            cell.border = thin_border
                            if cell.row > 1:
                                cell.alignment = Alignment(horizontal='left', vertical='center')
                    
                    # Add summary at bottom
                    summary_row = len(df_export) + 3
                    worksheet.cell(row=summary_row, column=1, value="Budget Summary")
                    worksheet.cell(row=summary_row, column=1).font = Font(bold=True, size=12)
                    
                    total = self.budget_df['Amount'].sum()
                    worksheet.cell(row=summary_row+1, column=1, value="Total Budget:")
                    worksheet.cell(row=summary_row+1, column=2, value=format_currency(total, self.currency))
                    
                    worksheet.cell(row=summary_row+2, column=1, value="Budget Type:")
                    worksheet.cell(row=summary_row+2, column=2, value=self.budget_type.upper() if self.budget_type else 'Custom')
                    
                    worksheet.cell(row=summary_row+3, column=1, value="Currency:")
                    worksheet.cell(row=summary_row+3, column=2, value=self.currency)
                    
                    worksheet.cell(row=summary_row+4, column=1, value="Created:")
                    worksheet.cell(row=summary_row+4, column=2, value=datetime.now().strftime('%Y-%m-%d %H:%M:%S'))
                
                print(f"✓ Excel report saved: {excel_path}")
                saved_files.append(excel_path)
                
            except Exception as e:
                print(f"⚠️  Excel export failed: {str(e)}")
                print("Please ensure openpyxl is installed: pip install openpyxl")
        
        return saved_files[0] if saved_files else None
    
    def run_approval_process(self, csv_file_path):
        """Run approval process on created budget"""
        if not APPROVAL_SYSTEM_AVAILABLE:
            print("\n⚠️  Approval system not available")
            return
        
        print("\n" + "="*60)
        print("BUDGET APPROVAL PROCESS")
        print("="*60)
        
        run_approval = input("Would you like to run the approval process? (y/n): ").strip().lower()
        
        if run_approval in ['y', 'yes']:
            try:
                approval_system = BudgetApprovalSystem()
                approval_system.process_budget_approval(csv_file_path)
            except Exception as e:
                print(f"Error in approval process: {str(e)}")


def main():
    """Main comprehensive budget creation workflow"""
    print("="*70)
    print("COMPREHENSIVE BUDGET CREATION SYSTEM")
    print("="*70)
    print("Create professional budgets with templates, analysis, and reports")
    print()
    
    creator = ComprehensiveBudgetCreator()
    
    # Step 1: Select currency
    creator.select_currency()
    
    # Step 2: Select budget type
    use_template = creator.select_budget_type()
    
    # Step 3: Fill template or add custom items
    if use_template:
        if not creator.use_template():
            print("\n⚠️  Template filling failed. Starting custom entry...")
            creator.add_custom_items()
    else:
        creator.add_custom_items()
    
    # Check if we have items
    if not creator.budget_items:
        print("\n⚠️  No budget items created. Exiting.")
        return
    
    # Step 4: Create DataFrame
    creator.create_budget_dataframe()
    
    # Step 5: Display summary
    creator.display_budget_summary()
    
    # Step 6: Optional quarterly breakdown
    quarterly_choice = input("\nCreate quarterly breakdown? (y/n): ").strip().lower()
    if quarterly_choice in ['y', 'yes']:
        quarterly_df = creator.create_quarterly_breakdown()
        if quarterly_df is not None:
            creator.budget_df = quarterly_df
    
    # Step 7: Save budget
    saved_file = creator.save_budget()
    
    # Step 8: Optional approval
    if saved_file and saved_file.endswith('.csv') and APPROVAL_SYSTEM_AVAILABLE:
        creator.run_approval_process(saved_file)
    
    print("\n" + "="*70)
    print("✓ BUDGET CREATION COMPLETE")
    print("="*70)
    print(f"Files saved in: {creator.output_dir}/")
    print("\nNext steps:")
    print("- Review your budget report")
    print("- Run approval process if not done: python budget_approval.py")
    print("- Use complete workflow: python run_workflow.py")


if __name__ == "__main__":
    main()
